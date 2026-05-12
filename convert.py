#!/usr/bin/env python3
"""
convert.py — Rex Yap Property Listings Converter
将 listing_manager.xlsx 转换为 listingsData.js

列结构 (listing_manager.xlsx):
  ★ 输入列 (你填写的):
    A(1)   状态 Status        — active / inactive
    B(2)   徽章 Badge         — 下拉选择
    C(3)   图片文件名 Image    — 逗号分隔多图
    D(4)   楼名·地点 Name     — 主要名称
    E(5)   房间数 Bedrooms
    F(6)   卫生间数 Bathrooms
    G(7)   面积 sqft
    H(8)   娱乐设施 Facilities
    I(9)   保安人员 Guard Type
    J(10)  CCTV 监控
    K(11)  门禁 Access Level
    L(12)  价格 RM (数字)
    M(13)  备注 Remarks

  ⚡ 自动生成列 (有内容则优先使用，否则自动生成):
    N(14)  ⚡ Icon
    O(15)  ⚡ 标题 ZH      P(16) ⚡ Title EN    Q(17) ⚡ Title MS
    R(18)  ⚡ 特点 ZH      S(19) ⚡ Features EN  T(20) ⚡ Feat MS
    U(21)  ⚡ 保安摘要 ZH  (仅参考，三语均自动生成)
    V(22)  ⚡ 价格 ZH      W(23) ⚡ Price EN     X(24) ⚡ Price MS
    Y(25)  ⚡ WA ZH        Z(26) ⚡ WA EN        AB(28) ⚡ WA MS

Usage:
    python convert.py
    python convert.py --excel listing_manager.xlsx --output listingsData.js
"""

import json, sys, argparse
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("❌ 请先安装: pip install openpyxl")
    sys.exit(1)

# ── 输入列 (1-based) ──────────────────────────────────────────────────────────
COL_STATUS  = 1
COL_BADGE   = 2
COL_IMAGE   = 3
COL_NAME    = 4
COL_BEDS    = 5
COL_BATHS   = 6
COL_SQFT    = 7
COL_FACIL   = 8
COL_GUARD   = 9
COL_CCTV    = 10
COL_ACCESS  = 11
COL_PRICE   = 12
COL_REMARKS = 13

# ── ⚡ 自动生成列 ──────────────────────────────────────────────────────────────
COL_ICON     = 14
COL_TITLE_ZH = 15
COL_TITLE_EN = 16
COL_TITLE_MS = 17
COL_FEAT_ZH  = 18
COL_FEAT_EN  = 19
COL_FEAT_MS  = 20
COL_SEC_ZH   = 21  # 保安摘要 ZH（仅参考）
COL_PRICE_ZH = 22
COL_PRICE_EN = 23
COL_PRICE_MS = 24
COL_WA_ZH    = 25
COL_WA_EN    = 26
COL_WA_MS    = 28

DATA_START_ROW = 4

# ── Badge → icon ──────────────────────────────────────────────────────────────
BADGE_ICON = {
    "🏢 Condo 公寓":        "fa-building",
    "🏠 Apartment":         "fa-home",
    "🏡 Terrace 排屋":      "fa-house",
    "🏘️ Semi-D":           "fa-home",
    "🏰 Bungalow 独立洋房":  "fa-home",
    "🏬 Shop Lot 商铺":     "fa-store",
    "🏗️ New Launch 新盘":   "fa-drafting-compass",
    "🏢 SOHO":              "fa-city",
}

# ── 城市名三语对照 ─────────────────────────────────────────────────────────────
CITY_TRANS = {
    "亚庇": ("Kota Kinabalu", "Kota Kinabalu"),
    "斗湖": ("Tawau", "Tawau"),
    "山打根": ("Sandakan", "Sandakan"),
    "拿笃": ("Lahad Datu", "Lahad Datu"),
    "吧巴": ("Papar", "Papar"),
    "根地咬": ("Keningau", "Keningau"),
    "兰脑": ("Ranau", "Ranau"),
    "古达": ("Kudat", "Kudat"),
    "保佛": ("Beaufort", "Beaufort"),
    "KKIP": ("KKIP", "KKIP"),
}

# ── 房型关键词三语对照 ─────────────────────────────────────────────────────────
TYPE_TRANS = {
    "公寓":  ("Condo", "Kondominium"),
    "排屋":  ("Terrace House", "Rumah Teres"),
    "洋房":  ("Bungalow", "Banglo"),
    "商铺":  ("Shop Lot", "Lot Kedai"),
    "新楼盘": ("New Launch", "Projek Baru"),
    "新盘":  ("New Launch", "Projek Baru"),
}

# ── 娱乐设施词库 ──────────────────────────────────────────────────────────────
# 格式：中文关键词 → (English, Bahasa Melayu)
FACILITY_TRANS = {
    # 水上设施
    "游泳池":       ("Swimming Pool", "Kolam Renang"),
    "儿童泳池":     ("Children's Pool", "Kolam Kanak-Kanak"),
    "无边泳池":     ("Infinity Pool", "Kolam Renang Infinity"),
    "水上乐园":     ("Water Park", "Taman Air"),
    "按摩池":       ("Jacuzzi", "Jakuzi"),
    "水疗":         ("Spa", "Spa"),
    # 健身 & 运动
    "健身室":       ("Gymnasium", "Gim"),
    "健身房":       ("Gymnasium", "Gim"),
    "瑜伽室":       ("Yoga Room", "Bilik Yoga"),
    "羽毛球场":     ("Badminton Court", "Gelanggang Badminton"),
    "网球场":       ("Tennis Court", "Gelanggang Tenis"),
    "篮球场":       ("Basketball Court", "Gelanggang Bola Keranjang"),
    "足球场":       ("Football Field", "Padang Bola Sepak"),
    "慢跑径":       ("Jogging Track", "Laluan Berjoging"),
    "跑步道":       ("Running Track", "Laluan Berlari"),
    "多用途球场":   ("Multi-Purpose Court", "Gelanggang Pelbagai Guna"),
    # 休闲 & 社交
    "BBQ区":        ("BBQ Area", "Kawasan BBQ"),
    "BBQ 区":       ("BBQ Area", "Kawasan BBQ"),
    "烧烤区":       ("BBQ Area", "Kawasan BBQ"),
    "会所":         ("Club House", "Dewan Kelab"),
    "会客室":       ("Function Room", "Bilik Fungsi"),
    "宴会厅":       ("Banquet Hall", "Dewan Banquet"),
    "空中花园":     ("Sky Garden", "Taman Langit"),
    "花园":         ("Garden", "Taman"),
    "屋顶花园":     ("Rooftop Garden", "Taman Bumbung"),
    "观景台":       ("Observation Deck", "Dek Pemerhatian"),
    "露台":         ("Terrace", "Teres"),
    "休闲区":       ("Recreational Area", "Kawasan Rekreasi"),
    "儿童游乐场":   ("Children's Playground", "Taman Permainan Kanak-Kanak"),
    "游乐场":       ("Playground", "Taman Permainan"),
    "阅读室":       ("Reading Room", "Bilik Bacaan"),
    "图书馆":       ("Library", "Perpustakaan"),
    # 实用设施
    "停车场":       ("Parking", "Tempat Letak Kereta"),
    "有盖停车":     ("Covered Parking", "Tempat Letak Kereta Berteduh"),
    "访客停车":     ("Visitor Parking", "Tempat Letak Kereta Pelawat"),
    "电动车充电":   ("EV Charging", "Pengecasan EV"),
    "洗衣房":       ("Laundry Room", "Bilik Dobi"),
    "储藏室":       ("Storage Room", "Bilik Stor"),
    "垃圾收集":     ("Waste Collection", "Pengumpulan Sampah"),
    "高速网络":     ("High-Speed Internet", "Internet Berkelajuan Tinggi"),
    "光纤网络":     ("Fibre Internet", "Internet Gentian Optik"),
    # 商业 & 服务
    "便利店":       ("Mini Mart", "Mini Mart"),
    "咖啡厅":       ("Café", "Kafe"),
    "餐厅":         ("Restaurant", "Restoran"),
    "商业中心":     ("Commercial Centre", "Pusat Komersial"),
    "迷你超市":     ("Mini Supermarket", "Mini Pasar Raya"),
    "礼宾服务":     ("Concierge Service", "Perkhidmatan Concierge"),
}

# ── 保安词汇三语词库 ───────────────────────────────────────────────────────────
# 格式：中文关键词 → (English, Bahasa Melayu)
SECURITY_TRANS = {
    # 保安人员
    "24小时保安":   ("24-Hour Security Guard", "Pengawal Keselamatan 24 Jam"),
    "24hr保安":    ("24-Hour Security Guard", "Pengawal Keselamatan 24 Jam"),
    "24 小时保安":  ("24-Hour Security Guard", "Pengawal Keselamatan 24 Jam"),
    "保安":        ("Security Guard", "Pengawal Keselamatan"),
    # CCTV
    "CCTV监控":    ("CCTV Surveillance", "Pengawasan CCTV"),
    "CCTV 监控":   ("CCTV Surveillance", "Pengawasan CCTV"),
    "CCTV":       ("CCTV", "CCTV"),
    "监控":        ("Surveillance", "Pengawasan"),
    # 门禁
    "等级3门禁":   ("Level 3 Access Card", "Kad Akses Tahap 3"),
    "等级2门禁":   ("Level 2 Access Card", "Kad Akses Tahap 2"),
    "等级1门禁":   ("Level 1 Access Card", "Kad Akses Tahap 1"),
    "门禁系统":    ("Access Card System", "Sistem Kad Akses"),
    "门禁":        ("Access Card", "Kad Akses"),
    "刷卡门禁":    ("Card Access", "Akses Kad"),
    # 其他安全设施
    "围栏社区":    ("Gated Community", "Komuniti Berkawal"),
    "围栏":        ("Gated", "Berkawal"),
    "有盖停车场":  ("Covered Parking", "Tempat Letak Kereta Berteduh"),
    "访客系统":    ("Visitor Management System", "Sistem Pengurusan Pelawat"),
    "对讲机":      ("Intercom", "Interkom"),
    "智能锁":      ("Smart Lock", "Kunci Pintar"),
}

SHEETS = {
    "rentals":     "🏠 出租房源 Rentals",
    "forSale":     "🏡 出售房源 For Sale",
    "newLaunches": "🏗️ 新楼盘 New Launches",
}


def cv(ws, row, col):
    v = ws.cell(row=row, column=col).value
    s = str(v).strip() if v is not None else ""
    return "" if s in ("nan", "None") else s


def get_icon(badge):
    for key, icon in BADGE_ICON.items():
        if key in badge:
            return icon
    return "fa-home"


def build_badge(badge_raw, remarks):
    if remarks and len(remarks) < 25:
        return f"✨ {remarks}"
    return badge_raw or "🏠 房源"


def translate_vocab(text, vocab_dict, lang):
    """通用词汇翻译：将中文词汇替换为英文或马来文。"""
    if not text:
        return text
    idx = 0 if lang == "en" else 1  # (en, ms) tuple index
    result = text
    for zh, trans in sorted(vocab_dict.items(), key=lambda x: -len(x[0])):
        if zh in result:
            result = result.replace(zh, trans[idx])
    return result


def translate_facilities(text, lang):
    """翻译娱乐设施（支持 · 、 , 分隔多项）。"""
    if not text or lang == "zh":
        return text
    for sep in ["·", "、", ","]:
        if sep in text:
            parts = [translate_vocab(p.strip(), FACILITY_TRANS, lang) for p in text.split(sep)]
            return " · ".join(parts)
    return translate_vocab(text, FACILITY_TRANS, lang)


def translate_security(text, lang):
    """翻译保安词汇。"""
    if not text or lang == "zh":
        return text
    return translate_vocab(text, SECURITY_TRANS, lang)


def auto_translate_title(name_zh):
    """自动将楼名翻译为英文/马来文（识别城市名和房型）。"""
    name_en, name_ms = name_zh, name_zh
    for zh, (en, ms) in CITY_TRANS.items():
        if zh in name_zh:
            name_en = name_en.replace(zh, en)
            name_ms = name_ms.replace(zh, ms)
    for zh, (en, ms) in TYPE_TRANS.items():
        if zh in name_en:
            name_en = name_en.replace(zh, en)
        if zh in name_ms:
            name_ms = name_ms.replace(zh, ms)
    return name_en, name_ms


def build_features(beds, baths, sqft, facil, guard, cctv, access, lang):
    """
    生成特点文字，保安词汇根据 lang 翻译。
    """
    if lang == "zh":
        parts = [f"📍 {beds}房{baths}卫"] if beds and baths else []
        if sqft:  parts.append(f"面积 {sqft} sqft")
        if facil: parts.append(facil)
        sec_parts = [p for p in [guard, cctv, access] if p]
    elif lang == "en":
        parts = [f"📍 {beds}BR {baths}BA"] if beds and baths else []
        if sqft:  parts.append(f"{sqft} sqft")
        if facil: parts.append(translate_facilities(facil, "en"))
        sec_parts = [translate_security(p, "en") for p in [guard, cctv, access] if p]
    else:  # ms
        parts = [f"📍 {beds}Bilik {baths}Bilik Air"] if beds and baths else []
        if sqft:  parts.append(f"{sqft} sqft")
        if facil: parts.append(translate_facilities(facil, "ms"))
        sec_parts = [translate_security(p, "ms") for p in [guard, cctv, access] if p]

    if sec_parts:
        parts.append(f"🔒 {' · '.join(sec_parts)}")
    return " · ".join(parts)


def build_price(price_rm, category, lang):
    try:
        price_str = f"{float(price_rm):,.0f}"
    except (ValueError, TypeError):
        price_str = ""

    if not price_str or category == "newLaunches":
        return {"zh": "🔥 欢迎咨询价格",
                "en": "🔥 Price Upon Request",
                "ms": "🔥 Harga atas Permintaan"}[lang]
    if category == "rentals":
        return {"zh": f"💰 月租 RM {price_str}",
                "en": f"💰 Monthly Rent RM {price_str}",
                "ms": f"💰 Sewa Bulanan RM {price_str}"}[lang]
    return {"zh": f"💰 售价 RM {price_str}",
            "en": f"💰 Selling Price RM {price_str}",
            "ms": f"💰 Harga Jualan RM {price_str}"}[lang]


def build_wa(titles, prices, beds, baths, category, lang):
    t = titles[lang]
    p = prices[lang]
    bed_zh = f"📍 {beds}房{baths}卫" if beds and baths else ""
    bed_en = f"📍 {beds}BR {baths}BA" if beds and baths else ""
    bed_ms = f"📍 {beds}Bilik {baths}Bilik Air" if beds and baths else ""

    if category == "rentals":
        return {
            "zh": f"你好Rex，我想了解出租房源：{t}（{p}）。\n{bed_zh}\n能否发送该房源的照片？谢谢！",
            "en": f"Hi Rex, I'm interested in the rental: {t} ({p}).\n{bed_en}\nCould you send photos? Thanks!",
            "ms": f"Hai Rex, saya berminat dengan sewa: {t} ({p}).\n{bed_ms}\nBoleh hantar gambar? Terima kasih!",
        }[lang]
    elif category == "newLaunches":
        return {
            "zh": f"你好Rex，我想了解新楼盘：{t}。\n能否发送项目照片/户型图？谢谢！",
            "en": f"Hi Rex, I'm interested in the new launch: {t}.\nCould you send project photos/floor plans? Thanks!",
            "ms": f"Hai Rex, saya berminat dengan projek baru: {t}.\nBoleh hantar gambar projek/pelan lantai? Terima kasih!",
        }[lang]
    else:
        return {
            "zh": f"你好Rex，我想了解出售房源：{t}（{p}）。\n{bed_zh}\n能否发送照片？谢谢！",
            "en": f"Hi Rex, I'm interested in the property: {t} ({p}).\n{bed_en}\nCould you send photos? Thanks!",
            "ms": f"Hai Rex, saya berminat dengan hartanah: {t} ({p}).\n{bed_ms}\nBoleh hantar gambar? Terima kasih!",
        }[lang]


def parse_sheet(ws, category):
    items = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        if cv(ws, row, COL_STATUS).lower() != "active":
            continue
        name = cv(ws, row, COL_NAME)
        if not name:
            continue

        badge_raw = cv(ws, row, COL_BADGE)
        image     = cv(ws, row, COL_IMAGE) or None
        beds      = cv(ws, row, COL_BEDS)
        baths     = cv(ws, row, COL_BATHS)
        sqft      = cv(ws, row, COL_SQFT)
        facil     = cv(ws, row, COL_FACIL)
        guard     = cv(ws, row, COL_GUARD)
        cctv      = cv(ws, row, COL_CCTV)
        access    = cv(ws, row, COL_ACCESS)
        price_rm  = cv(ws, row, COL_PRICE)
        remarks   = cv(ws, row, COL_REMARKS)

        # ── ⚡ 优先读取 Excel 已填写内容 ──────────────────────────────────────
        icon_pre     = cv(ws, row, COL_ICON)
        title_zh_pre = cv(ws, row, COL_TITLE_ZH)
        title_en_pre = cv(ws, row, COL_TITLE_EN)
        title_ms_pre = cv(ws, row, COL_TITLE_MS)
        feat_zh_pre  = cv(ws, row, COL_FEAT_ZH)
        feat_en_pre  = cv(ws, row, COL_FEAT_EN)
        feat_ms_pre  = cv(ws, row, COL_FEAT_MS)
        price_zh_pre = cv(ws, row, COL_PRICE_ZH)
        price_en_pre = cv(ws, row, COL_PRICE_EN)
        price_ms_pre = cv(ws, row, COL_PRICE_MS)
        wa_zh_pre    = cv(ws, row, COL_WA_ZH)
        wa_en_pre    = cv(ws, row, COL_WA_EN)
        wa_ms_pre    = cv(ws, row, COL_WA_MS)

        # ── 自动生成 ──────────────────────────────────────────────────────────
        icon  = icon_pre or get_icon(badge_raw)
        badge = build_badge(badge_raw, remarks)

        # 标题（三语）
        title_zh = title_zh_pre or name
        title_en_auto, title_ms_auto = auto_translate_title(name)
        title_en = title_en_pre or title_en_auto
        title_ms = title_ms_pre or title_ms_auto
        titles = {"zh": title_zh, "en": title_en, "ms": title_ms}

        # 特点（三语，保安词汇各语言翻译）
        feat_zh = feat_zh_pre or build_features(beds, baths, sqft, facil, guard, cctv, access, "zh")
        feat_en = feat_en_pre or build_features(beds, baths, sqft, facil, guard, cctv, access, "en")
        feat_ms = feat_ms_pre or build_features(beds, baths, sqft, facil, guard, cctv, access, "ms")

        # 价格（三语）
        price_zh = price_zh_pre or build_price(price_rm, category, "zh")
        price_en = price_en_pre or build_price(price_rm, category, "en")
        price_ms = price_ms_pre or build_price(price_rm, category, "ms")
        prices = {"zh": price_zh, "en": price_en, "ms": price_ms}

        # WhatsApp 文字（三语）
        wa_zh = wa_zh_pre or build_wa(titles, prices, beds, baths, category, "zh")
        wa_en = wa_en_pre or build_wa(titles, prices, beds, baths, category, "en")
        wa_ms = wa_ms_pre or build_wa(titles, prices, beds, baths, category, "ms")

        items.append({
            "badge":    badge,
            "icon":     icon,
            "image":    image,
            "title":    titles,
            "features": {"zh": feat_zh, "en": feat_en, "ms": feat_ms},
            "price":    prices,
            "waText":   {"zh": wa_zh, "en": wa_en, "ms": wa_ms},
        })
    return items


def main():
    parser = argparse.ArgumentParser(description="Convert listing_manager.xlsx → listingsData.js")
    parser.add_argument("--excel",  default="listing_manager.xlsx")
    parser.add_argument("--output", default="listingsData.js")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"❌ 找不到文件: {excel_path}")
        sys.exit(1)

    print(f"📖 读取: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    data, total = {}, 0
    for key, sheet_name in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  找不到工作表: '{sheet_name}' — 跳过")
            data[key] = []
            continue
        items = parse_sheet(wb[sheet_name], key)
        data[key] = items
        total += len(items)
        print(f"   ✅ {sheet_name}: {len(items)} 条")
        for it in items:
            print(f"      标题 ZH: {it['title']['zh']}")
            print(f"           EN: {it['title']['en']}")
            print(f"           MS: {it['title']['ms']}")
            print(f"      特点 ZH: {it['features']['zh']}")
            print(f"           EN: {it['features']['en']}")
            print(f"           MS: {it['features']['ms']}")

    js = (
        "// listingsData.js — 由 convert.py 自动生成，请勿手动编辑\n"
        "// Auto-generated by convert.py — DO NOT EDIT MANUALLY\n"
        f"// 最后更新 Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
        f"var listingsData = {json.dumps(data, ensure_ascii=False, indent=2)};\n"
    )

    Path(args.output).write_text(js, encoding="utf-8")
    print(f"\n✅ 完成！共 {total} 条房源 → {args.output}")
    print("💡 如需自定义翻译，可在 Excel ⚡ 列直接填写，优先级高于自动生成。")


if __name__ == "__main__":
    main()
